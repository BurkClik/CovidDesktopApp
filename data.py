import requests
from openpyxl import Workbook, load_workbook
import datetime
import os


def create_data_excel(name):
    """
    A function creates an excel file as `name.xlsx`

    Parameters
    ----------
    ``name`` : ``str``
        The name of excel file to be created.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Date"
    sheet["B1"] = "Country"
    sheet["C1"] = "Today Case"
    sheet["D1"] = "Today Deaths"
    sheet["E1"] = "Active Case"
    sheet["F1"] = "Tests"
    sheet["G1"] = "Recovered"
    sheet["H1"] = "Total Case"
    sheet["I1"] = "Total Deaths"

    workbook.save(filename="{}.xlsx".format(name))


def total_status_by_country(country):
    datas = []
    api = requests.get(
        'https://disease.sh/v2/countries/{country}?yesterday=false'.format(country=country))
    if api.json() != []:
        data = api.json()
        datas.append(data['todayCases'])
        datas.append(data['todayDeaths'])
        datas.append(data['active'])
        datas.append(data['tests'])
        datas.append(data['recovered'])
        datas.append(data['cases'])
        datas.append(data['deaths'])
        return datas


def total_status_by_country_excel(country):
    """
    A function shows the todayCases, todayDeaths,
    active, tests, recovered, cases,
    and deaths datas of ``country``

    Parameters
    ----------
    ``country`` : ``str``
        The country whose data are desired to be seen

    Return
    ------
    ``status{'Confirmed': confirmed_data, 'Deaths': deaths_data,
    'Recovered': recovered_data, 'Active': active_data}``
    """
    api = requests.get(
        'https://disease.sh/v2/countries/{country}?yesterday=false'.format(country=country.lower()))
    if api.json() != []:
        data = api.json()

        # If there is an excel file that has same name in the directory,
        # the code does not create an excel file. In this situation the 
        # code writes new datas below to old datas.
        directory = os.listdir()
        if '{}.xlsx'.format(country.capitalize()) not in directory:
            create_data_excel(country.capitalize())
            workbook = load_workbook('{}.xlsx'.format(country.capitalize()))
            sheet = workbook.active
            sheet['A2'] = datetime.date.today()
            sheet['B2'] = country.capitalize()
            sheet['C2'] = data['todayCases']
            sheet['D2'] = data['todayDeaths']
            sheet['E2'] = data['active']
            sheet['F2'] = data['tests']
            sheet['G2'] = data['recovered']
            sheet['H2'] = data['cases']
            sheet['I2'] = data['deaths']
            workbook.save(filename='{}.xlsx'.format(country.capitalize()))
            workbook.close()
        else:
            workbook = load_workbook('{}.xlsx'.format(country.capitalize()))
            sheet_range = workbook['Sheet']
            sheet = workbook.active
            i = 2
            while True:
                if sheet_range['B{}'.format(i)].value == None:
                    sheet['A{}'.format(i)] = datetime.date.today()
                    sheet['B{}'.format(i)] = country.capitalize()
                    sheet['C{}'.format(i)] = data['todayCases']
                    sheet['D{}'.format(i)] = data['todayDeaths']
                    sheet['E{}'.format(i)] = data['active']
                    sheet['F{}'.format(i)] = data['tests']
                    sheet['G{}'.format(i)] = data['recovered']
                    sheet['H{}'.format(i)] = data['cases']
                    sheet['I{}'.format(i)] = data['deaths']
                    break
                else:
                    i += 1
            workbook.save(filename='{}.xlsx'.format(country.capitalize()))
            workbook.close()
        return True
    else:
        return False

    
def just_global_excel():
    api = requests.get('https://disease.sh/v2/all?yesterday=false')
    data = api.json()

    directory = os.listdir()
    if 'global.xlsx' not in directory:
        create_data_excel("global")
        workbook = load_workbook(filename="global.xlsx")
        sheet = workbook.active
        sheet['A2'] = datetime.date.today()
        sheet['B2'] = "Global"
        sheet['C2'] = data['todayCases']
        sheet['D2'] = data['todayDeaths']
        sheet['E2'] = data['active']
        sheet['F2'] = data['tests']
        sheet['G2'] = data['recovered']
        sheet['H2'] = data['cases']
        sheet['I2'] = data['deaths']
        workbook.save(filename="global.xlsx")
        workbook.close()
    else:
        workbook = load_workbook(filename="global.xlsx")
        sheet_range = workbook['Sheet']
        sheet = workbook.active
        i = 2
        while True:
            if sheet_range['B{}'.format(i)].value == None:
                sheet['A{}'.format(i)] = datetime.date.today()
                sheet['B{}'.format(i)] = "Global"
                sheet['C{}'.format(i)] = data['todayCases']
                sheet['D{}'.format(i)] = data['todayDeaths']
                sheet['E{}'.format(i)] = data['active']
                sheet['F{}'.format(i)] = data['tests']
                sheet['G{}'.format(i)] = data['recovered']
                sheet['H{}'.format(i)] = data['cases']
                sheet['I{}'.format(i)] = data['deaths']
                break
            else:
                i += 1
        workbook.save(filename="global.xlsx")
        workbook.close()
        

def just_global():
    datas = []
    api = requests.get('https://disease.sh/v2/all?yesterday=false')
    data = api.json()
    datas.append(data['todayCases'])
    datas.append(data['todayDeaths'])
    datas.append(data['active'])
    datas.append(data['tests'])
    datas.append(data['recovered'])
    datas.append(data['cases'])
    datas.append(data['deaths'])
    return datas


def getCountries():
    api = requests.get('https://disease.sh/v2/countries?yesterday=false')
    data_countries = api.json()
    gcountries = []
    gcountries.append("Global")
    for i in data_countries:
        gcountries.append(i['country'])
    return gcountries


def sorted_data(sorted_by):
    """
    A function that pull sorted data.

    Parameters
    ----------
    ``sorted_by`` : ``str``
        The indicates which parameter
        will be pulled according to which parameter.
    """
    de = []
    api = requests.get('https://disease.sh/v2/countries?yesterday=false&sort={cases}'.format(cases=sorted_by))
    s_data = api.json()
    for i in s_data:
        de.append([i['country'], i['todayCases'], i['todayDeaths'],
                    i['active'], i['recovered'], i['cases'],
                    i['deaths']])
    return de